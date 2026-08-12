from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ---------------------------------------------------------
# 프로젝트 루트 설정
# ---------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


from app.db.connection import get_connection


# ---------------------------------------------------------
# Excel 저장 경로
# ---------------------------------------------------------

EXPORT_DIR = PROJECT_ROOT / "scripts" / "search"


# ---------------------------------------------------------
# CLI 인자
# ---------------------------------------------------------

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="kb.concepts 테이블을 DataFrame 형태로 조회합니다."
    )

    parser.add_argument(
        "--subject",
        type=str,
        default=None,
        help="subject_id 필터",
    )

    parser.add_argument(
        "--keyword",
        type=str,
        default=None,
        help="concept_id, name_ko, definition 검색",
    )

    parser.add_argument(
        "--limit",
        type=int,
        default=50,
        help="조회 최대 개수. 기본값: 50",
    )

    parser.add_argument(
        "--excel",
        type=str,
        default=None,
        help="Excel 파일명. 예: concepts.xlsx",
    )

    return parser.parse_args()


# ---------------------------------------------------------
# JSONB 배열 개수 SQL
# ---------------------------------------------------------

def json_array_count_sql(field_name: str) -> str:
    return f"""
        CASE
            WHEN jsonb_typeof(content -> '{field_name}') = 'array'
            THEN jsonb_array_length(content -> '{field_name}')
            ELSE 0
        END
    """


# ---------------------------------------------------------
# DB 조회
# ---------------------------------------------------------

def fetch_concepts(
    *,
    subject_id: str | None,
    keyword: str | None,
    limit: int,
) -> list[dict[str, Any]]:

    where_conditions: list[str] = []
    params: dict[str, Any] = {
        "limit": limit,
    }

    # subject 필터
    if subject_id:
        where_conditions.append(
            "subject_id = %(subject_id)s"
        )
        params["subject_id"] = subject_id

    # keyword 필터
    if keyword:
        where_conditions.append(
            """
            (
                concept_id ILIKE %(keyword)s
                OR name_ko ILIKE %(keyword)s
                OR content ->> 'definition' ILIKE %(keyword)s
            )
            """
        )

        params["keyword"] = f"%{keyword}%"

    if where_conditions:
        where_sql = (
            "WHERE "
            + " AND ".join(where_conditions)
        )
    else:
        where_sql = ""

    formula_count_sql = json_array_count_sql(
        "formulas"
    )

    condition_count_sql = json_array_count_sql(
        "application_conditions"
    )

    property_count_sql = json_array_count_sql(
        "properties"
    )

    prerequisite_count_sql = json_array_count_sql(
        "prerequisites"
    )

    related_count_sql = json_array_count_sql(
        "related_concepts"
    )

    objective_count_sql = json_array_count_sql(
        "learning_objectives"
    )

    misconception_count_sql = json_array_count_sql(
        "misconceptions"
    )

    sql = f"""
        SELECT
            concept_id,
            subject_id,
            chapter_id,
            section_id,
            name_ko,

            content ->> 'definition'
                AS definition,

            {formula_count_sql}
                AS formula_count,

            {condition_count_sql}
                AS condition_count,

            {property_count_sql}
                AS property_count,

            {prerequisite_count_sql}
                AS prerequisite_count,

            {related_count_sql}
                AS related_concept_count,

            {objective_count_sql}
                AS learning_objective_count,

            {misconception_count_sql}
                AS misconception_count,

            content
                -> 'generation_profile'
                ->> 'enabled'
                AS generation_enabled,

            content
                -> 'generation_profile'
                -> 'difficulty_range'
                ->> 'min'
                AS difficulty_min,

            content
                -> 'generation_profile'
                -> 'difficulty_range'
                ->> 'max'
                AS difficulty_max,

            content
                -> 'generation_profile'
                -> 'supported_problem_types'
                AS supported_problem_types,

            content
                -> 'generation_profile'
                -> 'supported_answer_types'
                AS supported_answer_types,

            content
                -> 'generation_profile'
                -> 'recommended_validators'
                AS recommended_validators,

            content
                -> 'generation_profile'
                ->> 'generation_notes'
                AS generation_notes,

            status,
            schema_version,
            content_version,
            created_at,
            updated_at

        FROM kb.concepts

        {where_sql}

        ORDER BY
            subject_id,
            chapter_id,
            concept_id

        LIMIT %(limit)s;
    """

    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                sql,
                params,
            )

            rows = cursor.fetchall()

        return rows

    finally:
        connection.close()


# ---------------------------------------------------------
# DataFrame 생성
# ---------------------------------------------------------

def create_dataframe(
    rows: list[dict[str, Any]],
) -> pd.DataFrame:

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    # definition 미리보기
    if "definition" in df.columns:
        df.insert(
            loc=df.columns.get_loc(
                "definition"
            ) + 1,
            column="definition_preview",
            value=df["definition"].apply(
                lambda value: (
                    value[:100] + "..."
                    if isinstance(value, str)
                    and len(value) > 100
                    else value
                )
            ),
        )

    return df


# ---------------------------------------------------------
# Excel 스타일 적용
# ---------------------------------------------------------

def style_excel(
    excel_path: Path,
) -> None:

    workbook = load_workbook(
        excel_path
    )

    worksheet = workbook[
        "Concepts"
    ]

    # 첫 행 고정
    worksheet.freeze_panes = "A2"

    # 자동 필터
    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    # Header 스타일
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # 기본 폭 계산
    for column_cells in (
        worksheet.columns
    ):
        column_letter = (
            column_cells[0].column_letter
        )

        max_length = 0

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            value_length = len(
                str(value)
            )

            if value_length > max_length:
                max_length = value_length

        width = min(
            max(max_length + 2, 10),
            40,
        )

        worksheet.column_dimensions[
            column_letter
        ].width = width

    # 특정 컬럼 넓게
    header_map = {
        cell.value: cell.column_letter
        for cell in worksheet[1]
    }

    if "definition" in header_map:
        worksheet.column_dimensions[
            header_map["definition"]
        ].width = 70

    if "definition_preview" in header_map:
        worksheet.column_dimensions[
            header_map["definition_preview"]
        ].width = 50

    if "generation_notes" in header_map:
        worksheet.column_dimensions[
            header_map["generation_notes"]
        ].width = 60

    if "supported_problem_types" in header_map:
        worksheet.column_dimensions[
            header_map[
                "supported_problem_types"
            ]
        ].width = 45

    if "supported_answer_types" in header_map:
        worksheet.column_dimensions[
            header_map[
                "supported_answer_types"
            ]
        ].width = 35

    if "recommended_validators" in header_map:
        worksheet.column_dimensions[
            header_map[
                "recommended_validators"
            ]
        ].width = 45

    # 모든 셀 줄바꿈
    for row in worksheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    # 행 높이
    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        worksheet.row_dimensions[
            row_number
        ].height = 45

    workbook.save(
        excel_path
    )


# ---------------------------------------------------------
# Excel 저장
# ---------------------------------------------------------

def save_excel(
    dataframe: pd.DataFrame,
    file_name: str,
) -> Path:

    EXPORT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    if not file_name.lower().endswith(
        ".xlsx"
    ):
        file_name += ".xlsx"

    excel_path = (
        EXPORT_DIR / file_name
    )

    dataframe.to_excel(
        excel_path,
        index=False,
        sheet_name="Concepts",
        engine="openpyxl",
    )

    style_excel(
        excel_path
    )

    return excel_path


# ---------------------------------------------------------
# 출력
# ---------------------------------------------------------

def print_dataframe(
    dataframe: pd.DataFrame,
) -> None:

    if dataframe.empty:
        print(
            "조회된 Concept가 없습니다."
        )
        return

    # 터미널 출력 폭 설정
    pd.set_option(
        "display.max_columns",
        None,
    )

    pd.set_option(
        "display.width",
        240,
    )

    pd.set_option(
        "display.max_colwidth",
        80,
    )

    print()
    print(
        dataframe.to_string(
            index=False
        )
    )


# ---------------------------------------------------------
# Main
# ---------------------------------------------------------

def main() -> None:

    args = parse_arguments()

    if args.limit <= 0:
        raise ValueError(
            "--limit 값은 1 이상이어야 합니다."
        )

    rows = fetch_concepts(
        subject_id=args.subject,
        keyword=args.keyword,
        limit=args.limit,
    )

    dataframe = create_dataframe(
        rows
    )

    print()
    print("=" * 70)
    print(
        f"조회된 Concept: "
        f"{len(dataframe)}개"
    )

    if args.subject:
        print(
            f"subject_id: "
            f"{args.subject}"
        )

    if args.keyword:
        print(
            f"검색어: "
            f"{args.keyword}"
        )

    print("=" * 70)

    print_dataframe(
        dataframe
    )

    if args.excel:

        if dataframe.empty:
            print()
            print(
                "조회 결과가 없어서 "
                "Excel 파일을 생성하지 않았습니다."
            )

        else:
            excel_path = save_excel(
                dataframe,
                args.excel,
            )

            print()
            print(
                "Excel 저장 완료:"
            )
            print(
                excel_path
            )


if __name__ == "__main__":
    main()